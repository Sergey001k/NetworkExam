import io

from openpyxl import Workbook
from openpyxl.styles import PatternFill


async def export_results(results, session_date):
    wb = Workbook()
    sheet = wb.active
    sheet.append(["ФИО", "Группа", "Количество баллов", "Результат (%)"])

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='E3F525', end_color='E3F525', fill_type='solid')
    green_fill = PatternFill(start_color='1FC425', end_color='1FC425', fill_type='solid')

    for result in results:
        sheet.append([result.student_name, result.group, result.score, f"{round(result.percent, 2)}%"])
        row = sheet.max_row
        percent_cell = sheet.cell(row=row, column=5)

        if result.percent < 60:
            percent_cell.fill = red_fill

        elif 60 <= result.percent <= 74:
            percent_cell.fill = yellow_fill

        else:
            percent_cell.fill = green_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Results-{session_date.date()} {session_date.hour}.{session_date.minute}.xlsx"

    return buffer, filename
